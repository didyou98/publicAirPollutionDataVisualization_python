# 라이브러리
import pandas as pd
import numpy as np
import openpyxl as xl
from datetime import date, datetime  # 날짜 제어
from dateutil.relativedelta import relativedelta  # datetime에서 지원하지 않는 메소드를 사용하기 위해 import
from openpyxl.styles import PatternFill # 각 셀의 배경색을 설정하기 위함

# 데이터 로드
df = pd.read_csv("공공데이터-2A-성연재-미세먼지.csv", header=0, index_col="날짜")

# 데이터 넣기(간단한 제어를 위해)
resion = {
    "서울": df["서울"],
    "부산": df["부산"],
    "대구": df["대구"],
    "인천": df["인천"],
    "광주": df["광주"],
    "대전": df["대전"],
    "울산": df["울산"],
    "경기": df["경기"],
    "강원": df["강원"],
    "충북": df["충북"],
    "충남": df["충남"],
    "전북": df["전북"],
    "전남": df["전남"],
    "경북": df["경북"],
    "경남": df["경남"],
    "제주": df["제주"],
    "세종": df["세종"]
}

# 미세먼지 평균 구하기
resionAvg = dict()
resionKeys = list(resion.keys())
for i in range(len(resion)):
    resionAvg[resionKeys[i]] = np.mean(resion[resionKeys[i]])
    print(resionAvg[resionKeys[i]])

# 각 지역의 max, min
resionMaxMin = dict()
for i in range(len(resion)):
    # idxmax()로 max를 가지는 지역을 찾는다
    # max()로 가장 큰 수치를 찾는다
    # 딕셔너리로 {"key", [v1, v2, v3, v4]} 형식을 갖는다
    resionMaxMin[resionKeys[i]] = [resion[resionKeys[i]].idxmax(), resion[resionKeys[i]].max(),
                                   resion[resionKeys[i]].idxmin(), resion[resionKeys[i]].min(skipna=True)]
maxResionName = ""
minResionName = ""
maxResionValue = 0
minResionValue = 999
reverseDic = {v: k for k, v in resionAvg.items()}
# 최대값, 최소값 찾기
for i in range(len(resionAvg)):
    if maxResionValue < resionAvg[resionKeys[i]]:
        maxResionValue = resionAvg[resionKeys[i]]
        maxResionName = reverseDic[maxResionValue]
    if minResionValue > resionAvg[resionKeys[i]]:
        minResionValue = resionAvg[resionKeys[i]]
        minResionName = reverseDic[minResionValue]
print(maxResionValue, maxResionName, minResionValue, minResionName)

filename = "(템플릿)출력데이터-2A-성연재-미세먼지.xlsx"
book = xl.load_workbook(filename)  # 템플릿 파일 로드
sheet = book["요약"]  # 시트 설정하고 액티브
sheet2 = book["기간별"]
sheet3 = book["기준별(한국)"]
sheet4 = book["기준별(WHO)"]
print(sheet3)
resionAvgList = list(resionAvg.values())

# 시트 1
# 평균치 넣기
for i in range(len(resionAvgList)):
    sheet[str(chr(66 + i)) + "10"] = resionAvgList[i]

# 지역과 평균 max, min 넣기
# str(chr(65)) = "A" + "15" 임으로 A15를 선택
sheet[str(chr(66)) + "15"] = maxResionName
sheet[str(chr(67)) + "15"] = maxResionValue
sheet[str(chr(66)) + "16"] = minResionName
sheet[str(chr(67)) + "16"] = minResionValue

# 각 지역 max, min, date 넣기
for i in range(len(resionAvgList)):
    sheet[str(chr(66 + i)) + "11"] = resionMaxMin[resionKeys[i]][0]
    sheet[str(chr(66 + i)) + "12"] = resionMaxMin[resionKeys[i]][1]
    sheet[str(chr(66 + i)) + "13"] = resionMaxMin[resionKeys[i]][2]
    sheet[str(chr(66 + i)) + "14"] = resionMaxMin[resionKeys[i]][3]

resionByPeriod = dict()
resionByPeriodList = list()
# 기간설정  relativedelta 라이브러리는 day까지 입력을 받아야해서 리스트로 생성
resionByPeriodDate = ["2019-10", "2019-11",
                      "2019-12", "2020-01",
                      "2020-02", "2020-03",
                      "2020-04", "2020-05",
                      "2020-06", "2020-07",
                      "2020-08", "2020-09",
                      "2020-10", "2020-11"
                      ]

# 변수를 각각 따로 반복해야해서 while문 사용
dateVar1 = 0
dateVar2 = 0

# timedelta(months=1) 이 사용할 수 없어서 relativedelt 라이브러리 메소드를 사용 date객체에 접근해 산술연산이 가능함
day_delta = relativedelta(months=1)

# 데이터의 시작일
start_date = "2019-10-01"
# 데이터의 마지막일
end_date = "2019-11-01"
for i in range(14):
    for j in range(17):
        # 각각의 평균을 월별로 넣기위해 구간설정
        resionByPeriodList.append(df[(df.index >= start_date) & (df.index < end_date)][resionKeys[j]].mean())
    # 시작일 date를 date라이브러리 객체로 변환 후 날짜를 더함 fromisoformat() 문자열을 date객체로 변환
    start_date = date.fromisoformat(start_date) + day_delta
    end_date = date.fromisoformat(end_date) + day_delta

    #  isoformat() 변환된 객체를 다시 문자열로 고침
    start_date = start_date.isoformat()
    end_date = end_date.isoformat()

# 각각의 데이터를 딕셔너리의 키와 값으로 넣음
# resionByPeriodList는 [0:17]까지는 10월 데이터 [17:34]까지는 11월 데이터가 들어가 있음.
while dateVar1 < 14:
    resionByPeriod[resionByPeriodDate[dateVar1]] = resionByPeriodList[dateVar2:dateVar2 + 17]
    dateVar1 += 1
    dateVar2 += 17

dateVar1 = 0
dateVar2 = 0
dateVar3 = 10
# 시트 2번에 넣기
while dateVar1 < 14:
    while dateVar2 < 17:
        sheet2[str(chr(66 + dateVar2)) + str(dateVar3)] = resionByPeriod[resionByPeriodDate[dateVar1]][dateVar2]
        dateVar2 += 1
    dateVar2 = 0
    dateVar1 += 1
    dateVar3 += 1
dateVar1 = 0
dateVar2 = 0
dateVar3 = 10
# 시트 3번 넣기
while dateVar1 < 14:
    while dateVar2 < 17:
        sheet3[str(chr(66 + dateVar2)) + str(dateVar3)] = resionByPeriod[resionByPeriodDate[dateVar1]][dateVar2]
        # 들어간 값이 30보다 낮을 경우 표기를 "좋음" 셀의 색깔을 연두색으로 바꾼다
        if sheet3[str(chr(66 + dateVar2)) + str(dateVar3)].value < 30:
            sheet3[str(chr(66 + dateVar2)) + str(dateVar3)] = "좋음"
            sheet3[str(chr(66 + dateVar2)) + str(dateVar3)].fill = PatternFill(fgColor="0061F0", fill_type="solid")
        elif sheet3[str(chr(66 + dateVar2)) + str(dateVar3)].value < 80:
            sheet3[str(chr(66 + dateVar2)) + str(dateVar3)] = "보통"
            sheet3[str(chr(66 + dateVar2)) + str(dateVar3)].fill = PatternFill(fgColor="92D050", fill_type="solid")
        elif sheet3[str(chr(66 + dateVar2)) + str(dateVar3)].value < 150:
            sheet3[str(chr(66 + dateVar2)) + str(dateVar3)] = "나쁨"
            sheet3[str(chr(66 + dateVar2)) + str(dateVar3)].fill = PatternFill(fgColor="FFD966", fill_type="solid")
        elif sheet3[str(chr(66 + dateVar2)) + str(dateVar3)].value > 151:
            sheet3[str(chr(66 + dateVar2)) + str(dateVar3)] = "매우나쁨"
            sheet3[str(chr(66 + dateVar2)) + str(dateVar3)].fill = PatternFill(fgColor="F4B084", fill_type="solid")
        dateVar2 += 1
    dateVar2 = 0
    dateVar1 += 1
    dateVar3 += 1

dateVar1 = 0
dateVar2 = 0
dateVar3 = 10

# 시트 4번 넣기
while dateVar1 < 14:
    while dateVar2 < 17:
        sheet4[str(chr(66 + dateVar2)) + str(dateVar3)] = resionByPeriod[resionByPeriodDate[dateVar1]][dateVar2]
        if sheet4[str(chr(66 + dateVar2)) + str(dateVar3)].value < 30:
            sheet4[str(chr(66 + dateVar2)) + str(dateVar3)] = "좋음"
            sheet4[str(chr(66 + dateVar2)) + str(dateVar3)].fill = PatternFill(fgColor="0061F0", fill_type="solid")
        elif sheet4[str(chr(66 + dateVar2)) + str(dateVar3)].value < 50:
            sheet4[str(chr(66 + dateVar2)) + str(dateVar3)] = "보통"
            sheet4[str(chr(66 + dateVar2)) + str(dateVar3)].fill = PatternFill(fgColor="92D050", fill_type="solid")
        elif sheet4[str(chr(66 + dateVar2)) + str(dateVar3)].value < 100:
            sheet4[str(chr(66 + dateVar2)) + str(dateVar3)] = "나쁨"
            sheet4[str(chr(66 + dateVar2)) + str(dateVar3)].fill = PatternFill(fgColor="FFD966", fill_type="solid")
        elif sheet4[str(chr(66 + dateVar2)) + str(dateVar3)].value > 101:
            sheet4[str(chr(66 + dateVar2)) + str(dateVar3)] = "매우나쁨"
            sheet4[str(chr(66 + dateVar2)) + str(dateVar3)].fill = PatternFill(fgColor="F4B084", fill_type="solid")
        dateVar2 += 1
    dateVar2 = 0
    dateVar1 += 1
    dateVar3 += 1

book.save("출력데이터-2A-성연재-미세먼지.xlsx")
print("ok")